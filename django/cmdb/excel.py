from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Protection, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from ast import literal_eval
import json
from .constants import FieldMapping, FieldType, ValidationType
from .models import ModelInstance, ValidationRules, ModelFields, ModelFieldMeta, Models
from .serializers import ModelInstanceSerializer
from rest_framework import serializers
from .utils import password_handler

import logging
logger = logging.getLogger(__name__)


class ExcelHandler:

    HEADER_FONT = Font(bold=True)
    REQUIRED_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

    @classmethod
    def _setup_instance_name_column(cls, sheet):
        """设置 instance_name 列（第一列）"""
        col_letter = 'A'

        # 第一行：字段名称
        cell = sheet[f'{col_letter}1']
        cell.value = "instance_name\r\n实例唯一标识"
        cell.font = cls.HEADER_FONT
        cell.fill = cls.REQUIRED_FILL
        cell.alignment = cls.CENTER_ALIGNMENT

        # 第二行：字段类型
        cell = sheet[f'{col_letter}2']
        cell.value = "string\r\n字符串"
        cell.alignment = cls.CENTER_ALIGNMENT

        # 第三行：约束说明
        cell = sheet[f'{col_letter}3']
        cell.value = "必填"
        cell.alignment = cls.CENTER_ALIGNMENT

        # 设置列宽
        sheet.column_dimensions[col_letter].width = 18

        # 设置数据验证
        dv = DataValidation(
            type='custom',
            formula1='AND(LEN(A4)>0,COUNTIF(A4:A1048576,A4)=1)',
            showErrorMessage=True,
            errorTitle='输入错误',
            error='实例名称不能为空且不能重复'
        )
        sheet.add_data_validation(dv)
        dv.add(f'{col_letter}4:{col_letter}1048576')

    @classmethod
    def _setup_using_template_column(cls, sheet):
        """设置 using_template 列（第二列）"""
        col_letter = 'B'

        # 第一行：字段名称
        cell = sheet[f'{col_letter}1']
        cell.value = "using_template\r\n是否使用模板自动重命名唯一标识"
        cell.font = cls.HEADER_FONT
        cell.alignment = cls.CENTER_ALIGNMENT

        # 第二行：字段类型
        cell = sheet[f'{col_letter}2']
        cell.value = "boolean\r\n布尔值"
        cell.alignment = cls.CENTER_ALIGNMENT

        # 第三行：约束说明
        cell = sheet[f'{col_letter}3']
        cell.value = "不填写默认为TRUE，默认值不会覆盖已经配置的标识"
        cell.alignment = cls.CENTER_ALIGNMENT

        # 设置列宽
        sheet.column_dimensions[col_letter].width = 20

        # 设置数据验证
        dv = DataValidation(
            type='list',
            formula1='"TRUE,FALSE"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle='输入错误',
            error='该字段只能输入 TRUE 或 FALSE'
        )
        sheet.add_data_validation(dv)
        dv.add(f'{col_letter}4:{col_letter}1048576')

    @classmethod
    def _setup_field_column(cls, template_sheet, enum_sheet, field, col_index, current_enum_col, enum_data=None, ref_data=None):
        """
        设置单个字段列的表头、类型、约束和数据验证。

        :param template_sheet: 配置数据工作表
        :param enum_sheet: 枚举类型可选值工作表
        :param field: 字段配置对象
        :param col_index: 列索引（从2开始，因为第1列是 instance_name）
        :param current_enum_col: 当前枚举列位置
        :param enum_data: 枚举数据 {key: label}，为 None 时从数据库获取
        :param ref_data: 引用数据 {instance_id: instance_name}，为 None 时从数据库获取
        :return: 更新后的枚举列位置
        """
        col_letter = get_column_letter(col_index)

        # 第一行：字段名称
        cell = template_sheet[f'{col_letter}1']
        cell.value = f"{field.name}\r\n{field.verbose_name}"
        cell.font = cls.HEADER_FONT
        cell.alignment = cls.CENTER_ALIGNMENT

        # 第二行：字段类型
        cell = template_sheet[f'{col_letter}2']
        field_verbose_type = FieldMapping.FIELD_TYPES.get(field.type, field.type)
        cell.value = f"{field.type}\r\n{field_verbose_type}"
        cell.alignment = cls.CENTER_ALIGNMENT

        # 第三行：字段约束
        constraints = []
        if field.required:
            constraints.append("必填")
            template_sheet[f'{col_letter}1'].fill = cls.REQUIRED_FILL

        # 处理不同字段类型的数据验证
        if field.type == FieldType.BOOLEAN:
            dv = DataValidation(
                type='list',
                formula1='"TRUE,FALSE"',
                allow_blank=not field.required,
                showErrorMessage=True,
                errorTitle='输入错误',
                error='该字段只能输入 TRUE 或 FALSE'
            )
            template_sheet.add_data_validation(dv)
            dv.add(f'{col_letter}4:{col_letter}1048576')

        elif field.type == FieldType.ENUM:
            # 获取枚举数据
            if enum_data is None and field.validation_rule:
                enum_data = json.loads(field.validation_rule.rule)

            if enum_data:
                current_enum_col = cls._write_enum_to_sheet(
                    enum_sheet, field, enum_data, current_enum_col,
                    template_sheet, col_letter, is_ref=False
                )
                constraints.append("枚举值")

        elif field.type == FieldType.MODEL_REF:
            # 获取引用数据
            if ref_data is None and field.ref_model:
                ref_instances = ModelInstance.objects.filter(
                    model=field.ref_model
                ).values_list('id', 'instance_name')
                ref_data = {str(id): instance_name for id, instance_name in ref_instances}

            if ref_data:
                current_enum_col = cls._write_enum_to_sheet(
                    enum_sheet, field, ref_data, current_enum_col,
                    template_sheet, col_letter, is_ref=True
                )
                constraints.append("关联模型")

        # 添加验证规则约束说明
        if field.validation_rule:
            rule = field.validation_rule
            if rule.type == ValidationType.RANGE:
                constraints.append(f"数值范围: {rule.rule.replace(',', ' ~ ')}")
            elif rule.type == ValidationType.REGEX:
                constraints.append(f"正则规则: {rule.rule}")

        cell = template_sheet[f'{col_letter}3']
        cell.value = "\r\n".join(constraints) if constraints else ""
        cell.alignment = cls.CENTER_ALIGNMENT

        # 设置列宽
        template_sheet.column_dimensions[col_letter].width = 20

        # 设置数据格式
        number_format = FieldMapping.TYPE_EXCEL_FORMATS.get(field.type)
        if number_format:
            column = template_sheet.column_dimensions[col_letter]
            column.number_format = number_format

        return current_enum_col

    @classmethod
    def _write_enum_to_sheet(cls, enum_sheet, field, data_dict, current_enum_col, template_sheet, col_letter, is_ref=False):
        """
        将枚举或引用数据写入枚举 sheet，并设置数据验证。

        :param enum_sheet: 枚举工作表
        :param field: 字段配置
        :param data_dict: {key: label} 或 {instance_id: instance_name}
        :param current_enum_col: 当前枚举列位置
        :param template_sheet: 模板工作表
        :param col_letter: 数据列字母
        :param is_ref: 是否为模型引用
        :return: 更新后的枚举列位置
        """
        key_col_letter = get_column_letter(current_enum_col)
        val_col_letter = get_column_letter(current_enum_col + 1)

        # 设置标题（合并单元格）
        enum_sheet.merge_cells(f'{key_col_letter}1:{val_col_letter}1')
        enum_sheet[f'{key_col_letter}1'] = f'{field.name}\r\n{field.verbose_name}'
        enum_sheet[f'{key_col_letter}1'].font = cls.HEADER_FONT
        enum_sheet[f'{key_col_letter}1'].alignment = cls.CENTER_ALIGNMENT

        # 设置列标识
        if is_ref:
            enum_sheet[f'{key_col_letter}2'] = 'ID(勿填)'
            enum_sheet[f'{val_col_letter}2'] = '实例名称'
        else:
            enum_sheet[f'{key_col_letter}2'] = '填写值'
            enum_sheet[f'{val_col_letter}2'] = '对应内容'

        # 写入数据
        for i, (key, value) in enumerate(data_dict.items(), 3):
            enum_sheet[f'{key_col_letter}{i}'] = key
            enum_sheet[f'{val_col_letter}{i}'] = value

        # 设置列宽
        enum_sheet.column_dimensions[key_col_letter].width = 15
        enum_sheet.column_dimensions[val_col_letter].width = 20

        # 设置数据验证（使用显示值列）
        if len(data_dict) > 0:
            dv = DataValidation(
                type='list',
                formula1=f'=枚举类型可选值!${val_col_letter}$3:${val_col_letter}${len(data_dict) + 2}',
                allow_blank=not field.required,
                showErrorMessage=True,
                errorTitle='输入错误',
                error='请从列表中选择一个值'
            )
        else:
            dv = DataValidation(
                type='custom',
                formula1='FALSE',
                allow_blank=not field.required,
                showErrorMessage=True,
                errorTitle='输入错误',
                error='该字段暂无可选值'
            )
        template_sheet.add_data_validation(dv)
        dv.add(f'{col_letter}4:{col_letter}1048576')

        return current_enum_col + 2

    @staticmethod
    def _handle_enum_data(enum_sheet, field, current_enum_col, template_sheet, col_letter, header_font):
        """处理枚举数据"""
        key_col_letter = get_column_letter(current_enum_col)
        val_col_letter = get_column_letter(current_enum_col + 1)

        # 设置标题
        enum_sheet.merge_cells(f'{key_col_letter}1:{val_col_letter}1')
        enum_sheet[f'{key_col_letter}1'] = f'{field.name}\r\n{field.verbose_name}'
        enum_sheet[f'{key_col_letter}1'].font = header_font
        enum_sheet[f'{key_col_letter}1'].alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )

        # 设置列标识
        enum_sheet[f'{key_col_letter}2'] = '填写值'
        enum_sheet[f'{val_col_letter}2'] = '对应内容'

        # 获取枚举数据
        enum_data = {}
        if field.type == FieldType.MODEL_REF:
            ref_instances = ModelInstance.objects.filter(
                model=field.ref_model
            ).values_list('id', 'instance_name')
            enum_data = {str(id): instance_name for id, instance_name in ref_instances}
        else:
            enum_data = json.loads(field.validation_rule.rule)

        # 写入数据
        for i, (key, value) in enumerate(enum_data.items(), 3):
            enum_sheet[f'{key_col_letter}{i}'] = key
            enum_sheet[f'{val_col_letter}{i}'] = value

        # 设置列宽和样式
        enum_sheet.column_dimensions[key_col_letter].width = 15
        enum_sheet.column_dimensions[val_col_letter].width = 15

        # 设置数据验证
        if len(enum_data) > 0:
            dv = DataValidation(
                type='list',
                formula1=f'=枚举类型可选值!${val_col_letter}$3:${val_col_letter}${len(enum_data) + 2}',
                allow_blank=not field.required,
                showErrorMessage=True,
                errorTitle='输入错误',
                error='请从列表中选择一个值'
            )
        else:
            dv = DataValidation(
                type='custom',
                formula1='FALSE',
                allow_blank=not field.required,
                showErrorMessage=True,
                errorTitle='输入错误',
                error='该字段暂无可选值'
            )
        template_sheet.add_data_validation(dv)
        dv.add(f'{col_letter}4:{col_letter}1048576')

        return current_enum_col + 2, "枚举值" if field.type != FieldType.MODEL_REF else "关联模型"

    @classmethod
    def _finalize_workbook(cls, template_sheet, enum_sheet):
        """完成工作簿的最终设置：冻结行、设置行高、锁定枚举表"""
        # 冻结前三行
        template_sheet.freeze_panes = 'A4'

        # 设置行高
        template_sheet.row_dimensions[1].height = 45
        template_sheet.row_dimensions[2].height = 30
        template_sheet.row_dimensions[3].height = 45
        enum_sheet.row_dimensions[1].height = 30

        # 锁定枚举表 sheet
        enum_sheet.protection.enable()
        enum_sheet.protection.set_password('123456')

        # 获取父工作簿并设置计算属性
        wb = template_sheet.parent
        wb.calculation.calcMode = 'auto'

    @staticmethod
    def generate_data_export(fields, instances):
        """生成实例数据导出"""
        # TODO: 导出时同时添加枚举类索引，方便修改，统一导出数据及模板格式
        # TODO: 将模板表头应用到数据导出中，抛弃原有的数据导出表头，方便再次导入
        wb = Workbook()
        data_sheet = wb.active
        data_sheet.title = "配置数据"

        # 复用模板格式设置
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')

        # 写入标题行
        col = 1
        name_col = get_column_letter(col)
        data_sheet[f'{name_col}1'] = "instance_name"
        data_sheet[f'{name_col}1'].font = header_font
        data_sheet[f'{name_col}1'].alignment = center_alignment

        # 写入字段标题
        for field in fields:
            col += 1
            col_letter = get_column_letter(col)
            data_sheet[f'{col_letter}1'] = f'{field.name}\r\n{field.verbose_name}'
            data_sheet[f'{col_letter}1'].font = header_font
            data_sheet[f'{col_letter}1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            data_sheet.column_dimensions[col_letter].width = 20

        field_to_col = {
            field.name: get_column_letter(idx + 2)
            for idx, field in enumerate(fields)
        }

        # 写入实例数据
        for row, instance in enumerate(instances, 2):
            # 写入name
            data_sheet[f'A{row}'] = instance.get('instance_name')

            field_values = instance.get('fields', {})

            # 写入字段值
            for field_name, field_value in field_values.items():
                if field_name in field_to_col:
                    col_letter = field_to_col[field_name]
                    data_sheet[f'{col_letter}{row}'] = field_value

        data_sheet.row_dimensions[1].height = 30
        data_sheet.freeze_panes = 'A2'

        return wb

    def validate_template(self, wb):
        try:
            if not {'配置数据', '枚举类型可选值'}.issubset(wb.sheetnames):
                raise serializers.ValidationError('Invalid file format. Missing required sheets.')
            sheet = wb['配置数据']
            if sheet['A1'].value != 'instance_name\n实例唯一标识' or \
                    sheet['A2'].value != 'string\n字符串' or \
                    sheet['A3'].value != '必填':
                raise serializers.ValidationError('Invalid file format. Missing required rows in template sheet.')
            return True
        except Exception as e:
            raise serializers.ValidationError(f"Invalid file format: {str(e)}")

    def get_cell_value(self, cell):
        if cell.value in ('#N/A', '#REF!', '#VALUE!', '#DIV/0!', '#NUM!', '#NAME?', '#NULL!'):
            return None
        return cell.value

    def load_data(self, file_path):
        """从Excel导入实例数据"""
        results = {
            'status': 'pending',
            'instances': [],
            'headers': [],
            'header_rows': [],
            'errors': [],
            'results': {
                'total': 0,
                'valid': 0,
                'invalid': 0
            }
        }
        try:
            # 加载Excel文件
            wb = load_workbook(file_path, data_only=True)
            self.validate_template(wb)

            data_sheet = wb['配置数据']
            # enum_sheet = wb['枚举类型可选值']

            # 保存前三行表头
            for row in range(1, 4):
                header_row = []
                for col in range(1, data_sheet.max_column + 1):
                    cell = data_sheet.cell(row=row, column=col)
                    header_row.append(cell.value)
                results['header_rows'].append(header_row)

            # 获取字段映射
            field_mapping = {}
            for col in range(3, data_sheet.max_column + 1):
                column_letter = get_column_letter(col)
                field_cell = data_sheet[f'{column_letter}1']

                if not field_cell.value:
                    continue

                # 解析字段名格式：name\nverbose_name
                try:
                    field_name = field_cell.value.split('\n')[0].strip()
                    if field_name:
                        field_mapping[column_letter] = field_name
                    results['headers'].append(field_name)
                except (AttributeError, IndexError):
                    logger.warning(f"Invalid field name format in column {column_letter}")
                    continue

            # 读取数据行
            for row in range(4, data_sheet.max_row + 1):
                try:
                    row_data = {}
                    for col_letter, field_name in field_mapping.items():
                        cell = data_sheet[f'{col_letter}{row}']
                        if cell.value is not None:
                            row_data[field_name] = cell.value

                    if row_data:
                        results['results']['total'] += 1
                        instance_data = {
                            'instance_name': data_sheet[f'A{row}'].value,
                            'using_template': data_sheet[f'B{row}'].value,
                            'fields': row_data
                        }
                        results['instances'].append(instance_data)
                        results['results']['valid'] += 1

                except Exception as e:
                    results['results']['invalid'] += 1
                    results['errors'].append(f"Load data failed at row {row}: {str(e)}")

            results['status'] = 'success'
            return results

        except Exception as e:
            results['errors'].append(f"Load data failed: {str(e)}")
            results['status'] = 'failed'
            return results

    @staticmethod
    def generate_error_export(headers, header_rows, error_data):
        """生成错误数据导出
        Args:
            model: 模型实例
            headers: 原始表头列表
            error_data: [{
                'instance_name': 'instance_name',
                'fields': {'field1': 'value1', ...},
                'error': 'error message'
            }]
        """
        # 生成基础表格
        wb = Workbook()
        sheet = wb.active
        sheet.title = "配置数据"

        # 复制前三行表头
        for row_idx, row_data in enumerate(header_rows, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                if row_idx == 1:
                    cell.font = Font(bold=True)
                sheet.column_dimensions[get_column_letter(col_idx)].width = 20

        # 添加错误信息列
        error_col = sheet.max_column + 1
        error_cell = sheet.cell(row=1, column=error_col, value='错误信息')
        error_cell.font = Font(bold=True)
        error_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 填充错误数据和错误信息
        for row, error in enumerate(error_data, 4):
            # 按headers顺序填充字段值
            fields = error.get('fields', {})

            sheet['A' + str(row)].value = error.get('instance_name')
            sheet['A' + str(row)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            sheet['B' + str(row)].value = error.get('using_template')
            sheet['B' + str(row)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for col_idx, header in enumerate(headers, 3):
                value = fields.get(header)
                if value:
                    cell = sheet.cell(row=row, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # 填充错误信息
            error_msg = str(error.get('error', ''))
            cell = sheet.cell(row=row, column=error_col, value=error_msg)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 调整错误列宽
        sheet.column_dimensions[get_column_letter(error_col)].width = 50

        return wb

    @classmethod
    def generate_template(cls, fields):
        """
        生成导入模板（仅表头，无数据）。

        :param fields: 字段配置列表
        :return: Workbook 对象
        """
        wb = Workbook()
        template_sheet = wb.active
        template_sheet.title = "配置数据"
        enum_sheet = wb.create_sheet("枚举类型可选值")

        # 设置 instance_name 列
        cls._setup_instance_name_column(template_sheet)
        # 设置 using_template 列
        cls._setup_using_template_column(template_sheet)

        # 设置各字段列
        current_enum_col = 1
        for col, field in enumerate(fields, 3):
            current_enum_col = cls._setup_field_column(
                template_sheet, enum_sheet, field, col, current_enum_col
            )

        # 完成工作簿设置
        cls._finalize_workbook(template_sheet, enum_sheet)

        return wb

    @classmethod
    def generate_data_export_with_template(cls, fields, instances_data, enum_data=None, ref_data=None):
        """
        生成带数据的导出文件，格式与导入模板完全一致。
        包含：配置数据 sheet（三行表头 + 数据行）、枚举类型可选值 sheet（带锁定）

        :param fields: 字段配置列表
        :param instances_data: 实例数据列表 [{'instance_name': ..., 'fields': {...}}]
        :param enum_data: 枚举数据 {field_name: {key: label}}，为 None 时从数据库获取
        :param ref_data: 引用数据 {field_name: {instance_id: instance_name}}，为 None 时从数据库获取
        :return: Workbook 对象
        """
        wb = Workbook()
        template_sheet = wb.active
        template_sheet.title = "配置数据"
        enum_sheet = wb.create_sheet("枚举类型可选值")

        # 设置 instance_name 列
        cls._setup_instance_name_column(template_sheet)
        # 设置 using_template 列
        cls._setup_using_template_column(template_sheet)

        # 设置各字段列（传入预加载的枚举/引用数据）
        current_enum_col = 1
        field_to_col = {}
        for col, field in enumerate(fields, 3):
            field_enum_data = enum_data.get(field.name) if enum_data else None
            field_ref_data = ref_data.get(field.name) if ref_data else None

            current_enum_col = cls._setup_field_column(
                template_sheet, enum_sheet, field, col, current_enum_col,
                enum_data=field_enum_data, ref_data=field_ref_data
            )
            field_to_col[field.name] = get_column_letter(col)

        # 写入实例数据（从第4行开始）
        for row, instance_data in enumerate(instances_data, 4):
            # 写入 instance_name
            template_sheet[f'A{row}'] = instance_data.get('instance_name')
            template_sheet[f'A{row}'].alignment = cls.CENTER_ALIGNMENT

            # 写入字段值
            fields_values = instance_data.get('fields', {})
            for field_name, field_value in fields_values.items():
                if field_name in field_to_col and field_value is not None:
                    col_letter = field_to_col[field_name]
                    template_sheet[f'{col_letter}{row}'] = field_value
                    template_sheet[f'{col_letter}{row}'].alignment = cls.CENTER_ALIGNMENT

        # 完成工作簿设置
        cls._finalize_workbook(template_sheet, enum_sheet)

        return wb
