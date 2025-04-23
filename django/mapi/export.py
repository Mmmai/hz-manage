from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font, PatternFill, Protection, Alignment
from rest_framework import serializers

class exportHandler:
    @staticmethod
    def get_template(colList):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "门户列表"
        # 样式定义
        header_font = Font(bold=True)
        required_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        center_alignment = Alignment(horizontal='center', vertical='center')
        for i in colList:

            sheet.append(i)
        return wb
    @staticmethod
    def get_portal(colList):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "门户列表"
        # 样式定义
        header_font = Font(bold=True)
        required_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        center_alignment = Alignment(horizontal='center', vertical='center')
        for i in colList:
            sheet.append(i)
        return wb        
    @staticmethod
    def load_data(filePath):
        # 加载Excel文件
        wb = load_workbook(filePath)
        try:
            if not {'门户列表'}.issubset(wb.sheetnames):
                raise serializers.ValidationError('Invalid file format. Missing required sheets.')
            # return True
        except Exception as e:
            raise serializers.ValidationError(f"Invalid file format: {str(e)}")
        # self.validate_template(wb)
        portal_sheet = wb["门户列表"]
        allRow = []
        for row in portal_sheet.iter_rows(min_row=2,max_col=7,values_only=True):
            # print(row)
            allRow.append(row)
        return allRow